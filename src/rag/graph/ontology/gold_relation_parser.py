"""
Gold Relation Parser (D-17)

READ-ONLY parser for the hand-authored gold-standard relationship triples
in column 22 (`graph_relation`) of the eval-report xlsx, sheet `eval-18`
(`src/results/evaluations/eval-report-hybrid-suite-20260630-0907.xlsx`).

Each cell is semi-structured PROSE with embedded triples, not strict
Turtle/JSON, e.g. (confirmed live against the actual cell content for
B01-001):

    (hospital_admin_system) -[SHARES_NETWORK_WITH]-> (CII);
    (hospital_admin_system) -[NOT DESIGNATED_AS]-> (CII);
    (Commissioner) -[DESIGNATES]-> (CII) [Cybersecurity_Act_2018 s7];
    (CCoP_2.0) -[APPLIES_TO]-> (designated_CII + cyber_operating_environment)
        [1.2.1, 1.4.1]

This is a regex-based EXTRACTOR, not a full parser — it pulls out the
`(subject) -[REL]-> (object)` triples and bracketed clause citations it can
confidently identify; free-form prose around/between them is ignored. The
script is READ-ONLY with respect to the input workbook — it produces a
coverage-gap report artifact for the Method-C human curation gate (D-14/D-17),
never mutates the xlsx.

Normalization pitfall (confirmed live): cell text uses `NOT DESIGNATED_AS`
(space) rather than the ontology's `NOT_DESIGNATED_AS` (underscore) —
`normalize_relation` collapses internal whitespace to underscores before any
set comparison against the ontology's relationship-type vocabulary.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Regex patterns (RESEARCH.md Q10, verified against live xlsx cell content)
# ---------------------------------------------------------------------------

TRIPLE_RE = re.compile(r"\(([^)]+)\)\s*-\[([^\]]+)\]->\s*\(([^)]+)\)")
CLAUSE_BRACKET_RE = re.compile(r"\[([\w.\s,()§]+)\]")

DEFAULT_SHEET_NAME = "eval-18"
COL_TEST_ID = 1
COL_GRAPH_RELATION = 22


def normalize_relation(raw: str) -> str:
    """Collapse internal whitespace in a raw relation label to underscores.

    "NOT DESIGNATED_AS" -> "NOT_DESIGNATED_AS". Leaves already-normalized
    labels (e.g. "SHARES_NETWORK_WITH") unchanged.
    """
    return re.sub(r"\s+", "_", raw.strip())


@dataclass
class CaseGoldRelations:
    """Parsed gold-relation content for one eval-18 test case (one xlsx row)."""

    test_id: str
    triples: list[tuple[str, str, str]] = field(default_factory=list)
    relation_types: set[str] = field(default_factory=set)
    entity_terms: set[str] = field(default_factory=set)
    clause_citations: list[str] = field(default_factory=list)


def parse_graph_relation_cell(
    cell_text: str,
) -> tuple[list[tuple[str, str, str]], set[str], set[str], list[str]]:
    """
    Parse one `graph_relation` cell into (triples, relation_types,
    entity_terms, clause_citations).

    Relation names are normalized (whitespace -> underscore) before being
    added to `relation_types` / triple tuples.

    Clause-citation bracket detection filters out relation-name brackets
    (e.g. the `[SHARES_NETWORK_WITH]` inside `-[SHARES_NETWORK_WITH]->`)
    which would otherwise also match CLAUSE_BRACKET_RE's character class
    (letters/underscore/whitespace overlaps a bare relation name). Real
    clause citations always contain at least one digit ("1.2.1",
    "Cybersecurity_Act_2018 s7", "RTF 2.2", "5.3.1(c)") while relation names
    never do — this is used as the disambiguating filter.
    """
    if not cell_text:
        return [], set(), set(), []

    raw_triples = TRIPLE_RE.findall(cell_text)
    triples: list[tuple[str, str, str]] = []
    relation_types: set[str] = set()
    entity_terms: set[str] = set()

    for subject, relation, obj in raw_triples:
        norm_relation = normalize_relation(relation)
        subject = subject.strip()
        obj = obj.strip()
        triples.append((subject, norm_relation, obj))
        relation_types.add(norm_relation)
        entity_terms.add(subject)
        entity_terms.add(obj)

    clause_citations: list[str] = []
    for group in CLAUSE_BRACKET_RE.findall(cell_text):
        if not any(ch.isdigit() for ch in group):
            # Relation-name bracket (e.g. "[SHARES_NETWORK_WITH]"), not a
            # clause citation — skip.
            continue
        for part in group.split(","):
            part = part.strip()
            if part:
                clause_citations.append(part)

    return triples, relation_types, entity_terms, clause_citations


def parse_gold_relations(
    xlsx_path: str | Path, sheet_name: str = DEFAULT_SHEET_NAME
) -> list[CaseGoldRelations]:
    """
    Read the `graph_relation` column from the eval-report xlsx and return
    one CaseGoldRelations per non-blank test_id row.
    """
    from openpyxl import load_workbook

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Gold-relation xlsx not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(
                f"Sheet {sheet_name!r} not found in {xlsx_path.name}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        cases: list[CaseGoldRelations] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue  # header row

            test_id = row[COL_TEST_ID - 1]
            if test_id is None or (isinstance(test_id, str) and not test_id.strip()):
                continue

            cell_text = row[COL_GRAPH_RELATION - 1] if len(row) >= COL_GRAPH_RELATION else None
            cell_text = str(cell_text) if cell_text else ""

            triples, relation_types, entity_terms, clause_citations = parse_graph_relation_cell(
                cell_text
            )
            cases.append(
                CaseGoldRelations(
                    test_id=str(test_id).strip(),
                    triples=triples,
                    relation_types=relation_types,
                    entity_terms=entity_terms,
                    clause_citations=clause_citations,
                )
            )

        logger.info(f"Parsed gold relations for {len(cases)} cases from {xlsx_path.name}")
        return cases
    finally:
        wb.close()


def all_gold_relation_types(cases: list[CaseGoldRelations]) -> set[str]:
    """Union of relation_types across all parsed cases."""
    types: set[str] = set()
    for case in cases:
        types |= case.relation_types
    return types
