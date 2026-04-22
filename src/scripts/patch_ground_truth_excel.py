#!/usr/bin/env python3
"""Apply ground-truth citation remaps to the expert-validation Excel workbook.

Encodes the remap proposals from:
  .planning/phases/03.2-corpus-ground-truth-correctness/audit-remap-proposal.md

Design:
  - Per-cluster application via --cluster flag (B08 | B09 | B22 | B24 | B07_422
    | B03_117 | B02_564 | B05_523 | B21_EXEMPT) OR --all for every REMAP-ALL
    block at once.
  - Dry-run mode prints row-by-row proposed edits without touching the file.
  - Non-dry-run creates a timestamped .bak copy next to the Excel file BEFORE
    modifying, and fails fast if the backup cannot be written.
  - B24 uses a per-row mapping table (multi-citation column 8).
  - B21 "exempt" mode appends a sentinel marker to col 19 (Remarks) rather
    than modifying citations — preserves hallucination-benchmark intent.
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[2]
EXCEL_PATH = (
    REPO_ROOT
    / "ground-truth"
    / "expert-validation"
    / "CCoP_V2_Test_Cases_Expert_Review.xlsx"
)
SHEET_NAME = "Test Cases Review"

# Column indices (1-based) as produced by generate_v2_expert_review.py
COL_TEST_ID = 1
COL_SECTION = 7
COL_CLAUSE_REFS = 8
COL_EXPECTED_RESP = 11
COL_REMARKS = 19


# ---- Remap rules ------------------------------------------------------------
# Each rule returns (new_section, new_clause_refs, expected_response_patch)
# where expected_response_patch is an optional (old_substring, new_substring)
# tuple, or None if no in-text change needed.


def _b08_rule(_test_id: str) -> tuple[str, str, tuple[str, str] | None]:
    return (
        "3",
        "3.2.2(b), 3.2.2(c) [support: Risk Assessment Guide §4.2, §4.3]",
        (
            "CCoP 2.0 Sections 4.2 (risk assessment)",
            "CCoP 2.0 Section 3.2.2 (risk assessment methodology); "
            "Risk Assessment Guide §4.2 Risk Analysis; §4.3 Risk Evaluation",
        ),
    )


def _b09_rule(_test_id: str) -> tuple[str, str, tuple[str, str] | None]:
    return (
        "3",
        "3.2.2(a), 3.2.4, 3.2.5 [support: Risk Assessment Guide §4.1]",
        (
            "4.2.1",
            "3.2.2(a) / 3.2.4 / 3.2.5",
        ),
    )


def _b22_rule(_test_id: str) -> tuple[str, str, tuple[str, str] | None]:
    return (
        "1",
        "1.6.1, 1.6.2, 1.6.3 [support: Cybersecurity Act 2018 §11(7)]",
        ("11.7", "1.6.1 / 1.6.2 / 1.6.3 (Cybersecurity Act §11(7))"),
    )


def _b07_422_rule(_test_id: str) -> tuple[str, str, tuple[str, str] | None]:
    return (
        "3",
        "3.2.2(b), 3.2.2(c) [support: Risk Assessment Guide §4.2, §4.3]",
        ("4.2.2", "3.2.2(b) / 3.2.2(c)"),
    )


def _b03_117_rule(_test_id: str) -> tuple[str, str, tuple[str, str] | None]:
    return (
        "1",
        "1.6.1, 1.6.2, 1.6.3 [support: Cybersecurity Act 2018 §11(7)]",
        ("11.7", "1.6.1 / 1.6.2 / 1.6.3 (Cybersecurity Act §11(7))"),
    )


def _b02_564_rule(_test_id: str) -> tuple[str, str, tuple[str, str] | None]:
    # Topic is patch management, not Network Security. §5.10.1(e) is the
    # "timely manner" patch clause. Per user decision (Phase B, Option 3),
    # the ER text is also rewritten to replace fabricated 14-day/30-day
    # timelines with CCoP's actual "timely manner" + risk-based prioritisation
    # language — handled by _B02_ER_REWRITE below via the full-ER mechanism.
    # This rule corrects the section/clause; the full ER replacement is
    # applied as a second pass.
    return (
        "5",
        "5.10.1(e)",
        ("5.6.4", "5.10.1(e)"),
    )


# Full-ER rewrites for B02_564 cluster. Replaces col 11 entirely for the
# four flagged rows so the expected response cites only CCoP 2.0 language
# that is verifiable against the source document. No fabricated timelines.
_B02_ER_REWRITE: dict[str, str] = {
    "B2-003": (
        "This configuration is compliant with CCoP 2.0 Clause 5.10.1(e), which "
        "requires security patches to be applied in a timely manner to reduce "
        "cybersecurity vulnerabilities, and Clause 5.10.1(d), which requires "
        "prioritising patch application based on the level of risk posed to the "
        "operations of the CII. A prompt deployment (within roughly two weeks) "
        "of a critical patch addressing a vulnerability with active exploitation "
        "or publicly available exploit code — including a short testing phase — "
        "reflects appropriate risk-based prioritisation and timely application. "
        "Clause 5.10.1(c) expects patches to be tested in an environment similar "
        "to the CII production environment before deployment; the testing "
        "period described does not compromise compliance. In the OT context, "
        "testing patches in an isolated lab before deploying to production SCADA "
        "is good practice. CCoP 2.0 does not prescribe a fixed number of days — "
        "'timely' is a risk-based determination — and the documented process "
        "meets that standard."
    ),
    "B2-010": (
        "This configuration is compliant with CCoP 2.0 Clause 5.10.1. Clause "
        "5.10.1(e) requires security patches to be applied in a timely manner; "
        "Clause 5.10.1(d) requires prioritisation based on the level of risk "
        "posed to operations. Applying critical patches (active exploitation or "
        "public exploit code) within about two weeks and non-critical patches "
        "within about three to four weeks is consistent with the timely, "
        "risk-prioritised application the clause expects. CCoP 2.0 does not "
        "prescribe a fixed number of days — timeliness is a risk-based "
        "determination. The two-track automated deployment process described "
        "also aligns with Clause 5.10.1(a) monitoring of patch releases, "
        "5.10.1(b) integrity verification, 5.10.1(c) testing prior to "
        "deployment, and 5.10.1(f) monitoring and tracking of patching progress."
    ),
    "B2-014": (
        "This is a partially compliant situation that likely requires a waiver. "
        "CCoP 2.0 Clause 5.10.1 requires security patches to be applied in a "
        "timely manner with prioritisation based on risk. When a vendor no "
        "longer supports equipment and issues no patches, there is no patch to "
        "apply — creating a genuine technical infeasibility. Clause 5.10.1(g) "
        "expects compensating controls to mitigate and reduce cybersecurity "
        "risks in cases where a security patch cannot be applied: network "
        "isolation, enhanced monitoring, and annual manual security review are "
        "appropriate responses to unpatched legacy OT equipment. However, "
        "operating end-of-life equipment with known unresolved vulnerabilities "
        "on CII is a compliance risk that the CIIO should address through the "
        "waiver process (Clause 1.6), which aligns with Section 11(7) of the "
        "Cybersecurity Act. The waiver application should document the genuine "
        "impossibility of patching, the 5.10.1(g) compensating controls in "
        "place, a timeline for equipment replacement or upgrade, and ongoing "
        "monitoring commitments. Compensating controls alone do not achieve "
        "full compliance with Clause 5.10 — a waiver formalises and legitimises "
        "the compensating control approach."
    ),
    "B2-024": (
        "This situation is non-compliant with CCoP 2.0 Clause 5.10.1(e) and "
        "Clause 5.10.1(d) for the specific critical patch. Clause 5.10.1(d) "
        "requires prioritising patch application based on the level of risk "
        "posed to operations; Clause 5.10.1(e) requires applying security "
        "patches in a timely manner to reduce cybersecurity vulnerabilities. "
        "Waiting for a fixed quarterly maintenance window before applying a "
        "critical patch for a vulnerability with active exploitation or a "
        "publicly available exploit does not reflect risk-based prioritisation "
        "or timeliness. A rigid quarterly maintenance schedule cannot override "
        "the Clause 5.10.1(d)/(e) obligation for critical patches. The "
        "organisation must either: (1) adjust the maintenance window schedule "
        "to allow out-of-cycle emergency patching for critical patches; (2) "
        "where that is genuinely infeasible, implement Clause 5.10.1(g) "
        "compensating controls (increased monitoring, network isolation) while "
        "pursuing emergency patching approval; or (3) if the patch cannot be "
        "applied in the near term, submit a waiver request under Clause 1.6, "
        "aligned with Section 11(7) of the Cybersecurity Act. Quarterly patch "
        "cycles may be acceptable for lower-risk non-critical patches when "
        "documented under Clause 5.10.1(d) risk-based prioritisation, but not "
        "for critical patches."
    ),
}


def _b05_523_rule(_test_id: str) -> tuple[str, str, tuple[str, str] | None]:
    # Topic is MFA (both B05-002 and B05-019). Matches the MFA singleton
    # bundle: §5.1.2 (generic auth), §5.3.1(c) (PAM MFA), §5.7.2(b) (remote MFA).
    return ("5", "5.1.2, 5.3.1, 5.7.2", ("Section 5.2.3", "Section 5.3.1(c)"))


# ---- B24 per-row table (primary citations from proposal §Cluster 4) --------

_B24_ROW_MAP: dict[str, str] = {
    "B24-001": "7.1.1(b), 7.1.1(g), 7.1.1(h) [support: 8.2.1]",
    "B24-002": "7.1.1(b), 7.1.1(g), 7.1.1(h) [support: 8.2.1]",
    "B24-003": "7.1.1(b), 7.1.1(g), 7.1.1(i), 7.1.4 [support: 8.2.1]",
    "B24-004": "7.1.1(d), 7.1.1(g) [support: 8.2.1]",
    "B24-005": "7.1.1(c), 7.1.1(d), 7.1.1(e)",
    "B24-006": "7.1.1(b), 7.1.1(f), 7.1.1(h)",
    "B24-007": "7.1.1(b), 7.1.1(g), 7.1.1(h)",
    "B24-008": "7.1.1(b), 7.1.1(g), 7.1.4",
    "B24-009": "7.1.1(b), 7.1.1(h), 7.1.1(i), 7.1.4",
    "B24-010": "7.1.1(g), 7.1.1(h)",
    "B24-011": "7.1.1(i), 7.1.4",
    "B24-012": "7.1.1(d), 7.1.1(g) [support: 8.2.1]",
    "B24-013": "7.1.1(i), 7.1.4",
    "B24-014": "7.1.1(d), 7.1.1(g) [support: 8.2.1]",
    "B24-015": "7.1.1(d), 7.1.1(g) [support: 8.2.1]",
    "B24-016": "7.1.1(g), 7.1.1(h)",
    "B24-017": "7.1.1(e), 7.1.1(g) [support: 7.2.2]",
    "B24-018": "7.1.1(g), 7.1.4",
    "B24-019": "7.1.1(d), 7.1.1(f)",
    "B24-020": "7.1.1(g) [support: 8.1.1, 8.1.2]",
    "B24-021": "7.1.1(g), 7.1.4 [support: 7.2.2]",
    # B24-022: BONUS finding — not in Pass-1 (existing 8.1, 8.2 refs resolve
    # in CCoP, but semantics are wrong: pre-incident threat-intel response).
    # User decision (Phase B, #4): include in Phase C with proper IR + threat
    # intelligence anchors.
    "B24-022": "6.4.1, 6.4.3, 7.1.1(a), 7.1.1(d) [support: 7.3.3(a)]",
    "B24-023": "7.1.1(g)",
    "B24-024": "7.1.1(c), 7.1.1(d), 7.1.1(g) [support: 8.2.1]",
    "B24-025": "7.1.1(b), 7.1.1(h), 7.1.1(i), 7.1.4",
}


# Per-row ER substitutions for selected B24 rows (multi-patch). Applied as
# a side-effect pass after the main B24 cluster so clause/section are
# already aligned. Each entry is a list of (old, new) substring pairs.
_B24_ER_PATCHES: dict[str, list[tuple[str, str]]] = {
    "B24-022": [
        (
            "Section 8.1 requires incident management policy include threat intelligence consumption",
            "Section 7.1 requires incident management to include threat-intelligence consumption (see Section 6.4)",
        ),
        (
            "Pre-incident preparation is Section 8.2 (IR plan)",
            "Pre-incident preparation is Section 7.1 (IR plan)",
        ),
        (
            "Section 5.1 (threat intelligence)",
            "Section 6.4 (threat intelligence)",
        ),
    ],
}


def _b24_rule(test_id: str) -> tuple[str, str, tuple[str, str] | None] | None:
    clause = _B24_ROW_MAP.get(test_id)
    if clause is None:
        return None
    return ("7", clause, None)


# ---- Singletons per-row table (from proposal "Singletons" section) ----------
# Each entry: (section, clause_refs, optional (old, new) in-text patch).

_SINGLETON_ROW_MAP: dict[str, tuple[str, str, tuple[str, str] | None]] = {
    "B1-001": (
        "1",
        "1.2.1, 1.4.1 [support: Cybersecurity Act 2018 §7, RESPONSE-TO-FEEDBACK Q2.2-2.3]",
        None,
    ),
    "B1-017": ("5", "5.1.2, 5.3.1, 5.7.2", ("Section 5.1.5", "Section 5.7.2")),
    "B2-001": ("5", "5.1.2, 5.7.2", ("Clause 5.1.5", "Clause 5.7.2(b)")),
    # ---- B3 full scope (27 rows, CCoP 2.0 verified; B3-004 & B3-011 handled
    # ---- by B03_117 cluster). Previously 5 entries (B3-005/006/019/021/024)
    # ---- were wrongly mapped to waiver — corrected below per actual ER topic.
    "B3-001": (
        "5",
        "5.2.1, 5.3.1(c)",
        None,
    ),
    "B3-002": (
        "5",
        "5.5.1, 5.5.2, 10.2.1",
        ("Section 5.4.1 explicitly requires", "Section 5.5 explicitly requires"),
    ),
    "B3-003": (
        "5",
        "5.2.1, 5.3.1(c)",
        ("CCoP 5.3.1 requires", "CCoP 5.2.1 / 5.3.1(c) requires"),
    ),
    "B3-005": (
        "5",
        "5.1.4",
        ("CCoP 5.3.2 requires", "CCoP 5.1.4 requires"),
    ),
    "B3-006": (
        "5",
        "5.10.1, 5.10.2",
        ("CCoP 4.2 requires systems be kept up to date", "CCoP 5.10 requires systems be kept up to date"),
    ),
    "B3-007": (
        "3",
        "3.8.1, 3.8.3, 3.8.4",
        ("additional obligations under CCoP Section 7", "additional obligations under CCoP Section 3.8"),
    ),
    "B3-008": (
        "3",
        "3.7.1, 3.7.3",
        ("CCoP Section 5 applies regardless of hosting model", "CCoP Section 3.7 applies regardless of hosting model"),
    ),
    "B3-009": (
        "3",
        "3.2.1, 3.2.5",
        ("CCoP Section 2 requires annual risk assessment", "CCoP Section 3.2 requires annual risk assessment"),
    ),
    "B3-010": (
        "3",
        "3.8.1, 3.8.2, 3.8.3",
        ("Outsourced monitoring can satisfy Section 5.2", "Outsourced monitoring can satisfy Section 3.8"),
    ),
    "B3-012": (
        "3",
        "3.1.1, 3.8.1",
        None,
    ),
    "B3-013": (
        "5",
        "5.3.1(c)",
        None,
    ),
    "B3-014": (
        "5",
        "5.3.1(a), 5.3.1(c)",
        None,
    ),
    "B3-015": (
        "3",
        "3.1.1, 3.3.1",
        None,
    ),
    "B3-016": (
        "5",
        "3.8.1, 5.7.1, 5.7.2",
        ("requires enhanced monitoring per CCoP Section 7", "requires enhanced monitoring per CCoP Sections 3.8 and 5.7"),
    ),
    "B3-017": (
        "6",
        "6.1.4",
        ("CCoP Section 5.2 requires log retention", "CCoP Section 6.1.4(c) requires log retention"),
    ),
    "B3-018": (
        "9",
        "9.2.1, 9.2.2",
        ("CCoP Section 6 requires role-specific training", "CCoP Section 9.2 requires role-specific training"),
    ),
    "B3-019": (
        "7",
        "7.3.1, 7.3.2",
        ("CCoP Section 8.5 requires incident response plans be tested", "CCoP Section 7.3 requires incident response plans be tested"),
    ),
    "B3-020": (
        "10",
        "10.2.1, 10.2.3",
        None,
    ),
    "B3-021": (
        "5",
        "5.10.1(d), 5.10.1(e), 5.10.1(g)",
        ("CCoP 4.2 requires systems be kept secure", "CCoP 5.10.1(g) allows compensating controls where a patch cannot be applied"),
    ),
    "B3-022": (
        "6",
        "6.4.1, 6.4.3",
        ("CCoP Section 5.1 requires risk-based security", "CCoP Section 6.4 requires threat-intelligence-driven security"),
    ),
    "B3-023": (
        "5",
        "5.3.1(c) [support: Cybersecurity Act 2018 §11(7)]",
        None,
    ),
    "B3-024": (
        "8",
        "8.1.4",
        ("CCoP Section 9.4 requires backup systems be tested", "CCoP Section 8.1.4 requires backup systems be tested"),
    ),
    "B3-025": (
        "5",
        "5.15.1, 5.15.3",
        ("does not fully satisfy CCoP 5.5", "does not fully satisfy CCoP 5.15"),
    ),
    "B3-026": (
        "4",
        "4.1.1, 4.1.2",
        ("CCoP Section 3 requires current asset inventory", "CCoP Section 4.1 requires current asset inventory"),
    ),
    "B3-027": (
        "3",
        "3.3.1, 3.3.3",
        ("CCoP Section 1 requires documented policies", "CCoP Section 3.3 requires documented policies"),
    ),
    "B3-028": (
        "3",
        "3.2.5 [support: 2.1.1]",
        None,
    ),
    "B3-029": (
        "1",
        "1.4.1, 5.1.2",
        None,
    ),
    "B3-030": (
        "1",
        "1.4.1, 1.4.5, 1.5.1",
        None,
    ),
    "B05-013": (
        "1",
        "1.6.1, 1.6.2, 1.6.3, 3.2.1 [support: Cybersecurity Act 2018 §11(7)]",
        ("Section 4.3", "Section 1.6 (Waiver)"),
    ),
    "B05-015": ("3", "3.8.1, 3.8.2, 3.8.3", ("Section 9.3.1", "Section 3.8")),
    "B05-016": (
        "5",
        "5.11.1, 5.11.2, 5.11.3, 5.11.4",
        ("Section 5.3.4", "Section 5.11"),
    ),
    # B05-018 handled by DEPRECATE cluster below, not here.
    "B06-002": ("5", "5.1.2, 5.3.1, 5.7.2", ("Section 5.2.3", "Section 5.3.1(c)")),
    "B06-013": ("5", "5.2.1, 5.2.2", ("Section 5.2.5", "Section 5.2.2")),
    "B06-018": ("8", "8.2.1, 8.2.2", ("Section 7.4.1", "Section 8.2")),
    "B06-019": (
        "3",
        "3.2.1, 3.2.2 [support: Risk Assessment Guide §3]",
        ("Section 4.2.1", "Section 3.2"),
    ),
    "B07-006": ("5", "5.2.1, 5.3.1", ("Section 5.2.4", "Section 5.2.1(c)")),
    "B07-007": ("5", "5.2.2, 5.3.1", ("Section 5.2.5", "Section 5.2.2")),
    "B07-008": ("5", "5.2.1, 5.3.1", ("Section 5.2.4", "Section 5.2.1(a)")),
    "B07-010": ("5", "5.3.1", ("Section 5.2.6", "Section 5.3.1")),
    "B07-015": ("6", "6.2.1, 6.2.2, 6.2.3", ("Section 6.3.4", "Section 6.2")),
    "B07-017": ("5", "5.5.1, 5.5.2, 10.2.1", ("Section 5.4.2", "Section 5.5")),
    "B07-018": ("5", "5.7.1, 5.7.2, 10.2.3", ("Section 5.4.4", "Section 5.7")),
    "B07-027": ("5", "5.1.2, 5.7.2", ("Section 5.2.3", "Section 5.7.2")),
    "B12-001": (
        "5",
        "5.1.2, 5.3.1, 5.7.2 [support: Auditing Guidelines for CII]",
        ("CCoP 2.0 5.2.3", "CCoP 2.0 §5.3.1(c)"),
    ),
    "B12-005": (
        "4",
        "4.1.1, 4.1.2 [support: Auditing Guidelines for CII]",
        ("CCoP 2.0 4.2.2", "CCoP 2.0 §4.1"),
    ),
    "B12-008": (
        "8",
        "8.2.1, 8.2.2, 8.2.3, 8.2.4 [support: Auditing Guidelines for CII]",
        ("CCoP 2.0 7.4.1", "CCoP 2.0 §8.2"),
    ),
    "B12-014": (
        "5",
        "5.2.1, 5.2.2 [support: Auditing Guidelines for CII]",
        ("CCoP 2.0 5.2.5", "CCoP 2.0 §5.2.2"),
    ),
    "B12-016": (
        "3",
        "3.8.1, 3.8.2, 3.8.3, 3.8.4, 3.8.5 [support: Auditing Guidelines for CII]",
        ("CCoP 2.0 9.3.1", "CCoP 2.0 §3.8"),
    ),
    "B12-020": (
        "3",
        "3.2.1, 3.2.2 [support: Auditing Guidelines for CII, Risk Assessment Guide §3]",
        ("CCoP 2.0 4.2.1", "CCoP 2.0 §3.2"),
    ),
}


def _singleton_rule(
    test_id: str,
) -> tuple[str, str, tuple[str, str] | None] | None:
    return _SINGLETON_ROW_MAP.get(test_id)


# ---- DEPRECATE cluster (B05-018: cross-border data out of CCoP scope) -------

_DEPRECATE_MAP: dict[str, str] = {
    "B05-018": "Cross-border data transfer scoped to PDPA, not CCoP 2.0",
}
DEPRECATE_MARKER_PREFIX = "[DEPRECATED: "


# ---- Cluster → (matcher, rule_fn, label) ------------------------------------

ClusterMatcher = Callable[[str, str], bool]  # (test_id, current_clause_cell) -> bool
RuleFn = Callable[[str], tuple[str, str, tuple[str, str] | None] | None]


def _bench_prefix(prefix: str) -> ClusterMatcher:
    return lambda tid, _cell: tid.startswith(prefix)


def _bench_and_clause(prefix: str, clause_needle: str) -> ClusterMatcher:
    return lambda tid, cell: tid.startswith(prefix) and clause_needle in (cell or "")


CLUSTERS: dict[str, tuple[ClusterMatcher, RuleFn, str]] = {
    "B08": (
        _bench_prefix("B08-"),
        _b08_rule,
        "B08 Risk Prioritization → 3.2.2(b)/3.2.2(c)",
    ),
    "B09": (
        _bench_prefix("B09-"),
        _b09_rule,
        "B09 Risk Identification → 3.2.2(a)/3.2.4/3.2.5",
    ),
    "B22": (
        _bench_prefix("B22-"),
        _b22_rule,
        "B22 Waiver → 1.6.1/1.6.2/1.6.3 + Act §11(7)",
    ),
    "B24": (
        # only rows present in the per-row map
        lambda tid, _cell: tid in _B24_ROW_MAP,
        _b24_rule,
        "B24 Incident Response → per-row 7.1.x mapping",
    ),
    "B07_422": (
        _bench_and_clause("B07-", "4.2.2"),
        _b07_422_rule,
        "B07 4.2.2 → 3.2.2(b)/3.2.2(c)",
    ),
    "B03_117": (
        _bench_and_clause("B3-", "11.7"),
        _b03_117_rule,
        "B03 11.7 → 1.6.1/1.6.2/1.6.3 + Act §11(7)",
    ),
    "B02_564": (
        _bench_and_clause("B2-", "5.6.4"),
        _b02_564_rule,
        "B02 5.6.4 → 5.10.1(e) (patch management; ER timeline claims unverified)",
    ),
    "B05_523": (
        _bench_and_clause("B05-", "5.2.3"),
        _b05_523_rule,
        "B05 5.2.3 → MFA bundle (5.1.2, 5.3.1, 5.7.2)",
    ),
    "SINGLETONS": (
        lambda tid, _cell: tid in _SINGLETON_ROW_MAP,
        _singleton_rule,
        "Verified singletons (52 rows incl. 28 B3 full-scope) — per-row mapping from proposal",
    ),
}

# B21 audit-exempt test IDs (from audit-review-worksheet §Executive Summary).
B21_EXEMPT_IDS = {
    "B21-001",
    "B21-004",
    "B21-005",
    "B21-008",
    "B21-009",
    "B21-010",
    "B21-012",
    "B21-016",
    "B21-018",
    "B21-019",
    "B21-021",
}
B21_EXEMPT_MARKER = "[AUDIT_EXEMPT: hallucination benchmark — by design]"

REMAP_ALL_CLUSTERS = ["B08", "B09", "B22", "B07_422", "B03_117", "B05_523"]


# ---- Core apply logic -------------------------------------------------------


def _backup(src: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    dst = src.with_suffix(src.suffix + f".{stamp}.bak")
    shutil.copy2(src, dst)
    return dst


def _apply_cluster(
    ws: Worksheet,
    cluster_key: str,
    dry_run: bool,
) -> tuple[int, int]:
    """Return (rows_matched, rows_modified)."""
    matcher, rule_fn, label = CLUSTERS[cluster_key]
    matched = 0
    modified = 0
    print(f"\n-- {label} --")
    for row in ws.iter_rows(min_row=2, values_only=False):
        test_id = row[COL_TEST_ID - 1].value
        if not test_id:
            continue
        current_clause = row[COL_CLAUSE_REFS - 1].value or ""
        if not matcher(str(test_id), str(current_clause)):
            continue
        matched += 1
        rule = rule_fn(str(test_id))
        if rule is None:
            print(f"  [skip] {test_id}: no rule entry")
            continue
        new_section, new_clause, expected_patch = rule

        section_cell = row[COL_SECTION - 1]
        clause_cell = row[COL_CLAUSE_REFS - 1]
        expected_cell = row[COL_EXPECTED_RESP - 1]

        section_changed = str(section_cell.value or "") != new_section
        clause_changed = str(clause_cell.value or "") != new_clause
        expected_changed = False
        if expected_patch is not None:
            old_sub, new_sub = expected_patch
            current_resp = str(expected_cell.value or "")
            if old_sub in current_resp:
                expected_changed = True

        if not (section_changed or clause_changed or expected_changed):
            print(f"  [noop] {test_id}: already at target")
            continue

        print(
            f"  [edit] {test_id}: "
            f"col7 '{section_cell.value}' → '{new_section}' | "
            f"col8 '{clause_cell.value}' → '{new_clause}'"
            + (" | col11 in-text patch" if expected_changed else "")
        )

        if not dry_run:
            if section_changed:
                section_cell.value = new_section
            if clause_changed:
                clause_cell.value = new_clause
            if expected_changed:
                old_sub, new_sub = expected_patch
                expected_cell.value = str(expected_cell.value).replace(old_sub, new_sub)
        modified += 1

    print(f"  total: matched={matched}, modified={modified}")
    return matched, modified


def _apply_b24_er_patches(ws: Worksheet, dry_run: bool) -> tuple[int, int]:
    """Apply multi-substring ER patches for selected B24 rows (e.g. B24-022).

    Runs after _apply_cluster("B24") has aligned section/clause. Each row
    in _B24_ER_PATCHES gets its list of (old, new) substring replacements
    applied sequentially. Idempotent: a patch whose `old` substring is not
    present is silently skipped.
    """
    matched = 0
    modified = 0
    print("\n-- B24 Expected-Response patches --")
    for row in ws.iter_rows(min_row=2, values_only=False):
        test_id = row[COL_TEST_ID - 1].value
        if not test_id or str(test_id) not in _B24_ER_PATCHES:
            continue
        matched += 1
        expected_cell = row[COL_EXPECTED_RESP - 1]
        current = str(expected_cell.value or "")
        updated = current
        applied: list[str] = []
        for old_sub, new_sub in _B24_ER_PATCHES[str(test_id)]:
            if old_sub in updated:
                updated = updated.replace(old_sub, new_sub)
                applied.append(old_sub[:40] + "...")
        if updated == current:
            print(f"  [noop] {test_id}: no ER substrings matched")
            continue
        print(f"  [edit] {test_id}: col11 patched ({len(applied)} substitutions)")
        if not dry_run:
            expected_cell.value = updated
        modified += 1
    print(f"  total: matched={matched}, modified={modified}")
    return matched, modified


def _apply_b02_er_rewrite(ws: Worksheet, dry_run: bool) -> tuple[int, int]:
    """Replace col 11 (Expected Response) wholesale for B02_564 cluster rows.

    Runs after _apply_cluster("B02_564") so section/clause/in-text patches
    are already applied. This step overwrites the full ER text to remove
    fabricated 14/30-day timelines and align with CCoP §5.10 "timely manner"
    language. Idempotent: skips rows whose current ER already matches the
    target text.
    """
    matched = 0
    modified = 0
    print("\n-- B02_564 Expected-Response rewrite (timely-manner language) --")
    for row in ws.iter_rows(min_row=2, values_only=False):
        test_id = row[COL_TEST_ID - 1].value
        if not test_id or str(test_id) not in _B02_ER_REWRITE:
            continue
        matched += 1
        expected_cell = row[COL_EXPECTED_RESP - 1]
        target = _B02_ER_REWRITE[str(test_id)]
        current = str(expected_cell.value or "")
        if current.strip() == target.strip():
            print(f"  [noop] {test_id}: ER already rewritten")
            continue
        print(f"  [edit] {test_id}: col11 full rewrite ({len(target)} chars)")
        if not dry_run:
            expected_cell.value = target
        modified += 1
    print(f"  total: matched={matched}, modified={modified}")
    return matched, modified


def _apply_deprecate(ws: Worksheet, dry_run: bool) -> tuple[int, int]:
    matched = 0
    modified = 0
    print("\n-- DEPRECATE marker --")
    for row in ws.iter_rows(min_row=2, values_only=False):
        test_id = row[COL_TEST_ID - 1].value
        if not test_id or str(test_id) not in _DEPRECATE_MAP:
            continue
        matched += 1
        remarks_cell = row[COL_REMARKS - 1]
        current = str(remarks_cell.value or "")
        reason = _DEPRECATE_MAP[str(test_id)]
        marker = f"{DEPRECATE_MARKER_PREFIX}{reason}]"
        if DEPRECATE_MARKER_PREFIX in current:
            print(f"  [noop] {test_id}: already marked deprecated")
            continue
        new_value = (current + " " + marker).strip()
        print(f"  [edit] {test_id}: col19 += '{marker}'")
        if not dry_run:
            remarks_cell.value = new_value
        modified += 1
    print(f"  total: matched={matched}, modified={modified}")
    return matched, modified


def _apply_b21_exempt(ws: Worksheet, dry_run: bool) -> tuple[int, int]:
    matched = 0
    modified = 0
    print("\n-- B21 audit-exempt marker --")
    for row in ws.iter_rows(min_row=2, values_only=False):
        test_id = row[COL_TEST_ID - 1].value
        if not test_id or str(test_id) not in B21_EXEMPT_IDS:
            continue
        matched += 1
        remarks_cell = row[COL_REMARKS - 1]
        current = str(remarks_cell.value or "")
        if B21_EXEMPT_MARKER in current:
            print(f"  [noop] {test_id}: already marked")
            continue
        new_value = (current + " " + B21_EXEMPT_MARKER).strip()
        print(f"  [edit] {test_id}: col19 += '{B21_EXEMPT_MARKER}'")
        if not dry_run:
            remarks_cell.value = new_value
        modified += 1
    print(f"  total: matched={matched}, modified={modified}")
    return matched, modified


# ---- CLI --------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--cluster",
        action="append",
        choices=list(CLUSTERS.keys()) + ["B21_EXEMPT", "DEPRECATE"],
        help="Cluster(s) to apply. Repeatable. DEPRECATE opts-in to B05-018 deprecation.",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help=f"Apply every REMAP-ALL cluster: {REMAP_ALL_CLUSTERS} + SINGLETONS + B21_EXEMPT. "
        "Does NOT apply B24 (per-row), B02/B05 provisional clusters, or DEPRECATE — use --cluster for those.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview changes without writing the file.",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=EXCEL_PATH,
        help="Path to expert-validation Excel (default: repo path).",
    )
    args = parser.parse_args()

    if not args.cluster and not args.all:
        parser.error("specify --cluster ... (repeatable) or --all")

    clusters_to_apply: list[str] = []
    if args.all:
        clusters_to_apply.extend(REMAP_ALL_CLUSTERS)
        clusters_to_apply.append("SINGLETONS")
        clusters_to_apply.append("B21_EXEMPT")
    if args.cluster:
        for c in args.cluster:
            if c not in clusters_to_apply:
                clusters_to_apply.append(c)

    if not args.excel.exists():
        print(f"ERROR: Excel not found at {args.excel}", file=sys.stderr)
        return 2

    if not args.dry_run:
        backup = _backup(args.excel)
        print(f"Backup written: {backup}")

    wb = load_workbook(args.excel)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        return 3
    ws = wb[SHEET_NAME]

    totals = {"matched": 0, "modified": 0}
    for key in clusters_to_apply:
        if key == "B21_EXEMPT":
            m, n = _apply_b21_exempt(ws, args.dry_run)
        elif key == "DEPRECATE":
            m, n = _apply_deprecate(ws, args.dry_run)
        else:
            m, n = _apply_cluster(ws, key, args.dry_run)
        totals["matched"] += m
        totals["modified"] += n
        # B02_564 has an ER-rewrite side-effect that runs after the cluster.
        if key == "B02_564":
            m2, n2 = _apply_b02_er_rewrite(ws, args.dry_run)
            totals["matched"] += m2
            totals["modified"] += n2
        # B24 has per-row ER patches (e.g. B24-022 BONUS finding).
        if key == "B24":
            m2, n2 = _apply_b24_er_patches(ws, args.dry_run)
            totals["matched"] += m2
            totals["modified"] += n2

    print("\n" + "=" * 60)
    print(f"CLUSTERS APPLIED: {clusters_to_apply}")
    print(f"TOTAL MATCHED:    {totals['matched']}")
    print(f"TOTAL MODIFIED:   {totals['modified']}")
    print(f"DRY RUN:          {args.dry_run}")
    print("=" * 60)

    if not args.dry_run:
        wb.save(args.excel)
        print(f"Wrote: {args.excel}")
    else:
        print("Dry run — no file modified.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
