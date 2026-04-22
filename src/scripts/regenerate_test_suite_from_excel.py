#!/usr/bin/env python3
"""Regenerate ground-truth/test-suite/*.jsonl from the expert-validation Excel.

After patch_ground_truth_excel.py applies remaps to the Excel workbook (the
authoritative source), this script syncs those corrections back into the
JSONL test-suite files consumed by the evaluation pipeline.

What gets synced per test_id:
  - metadata.section          ← Excel col 7 (CCoP Section)
  - metadata.clause_reference ← Excel col 8 primary clauses (parsed)
  - metadata.support_citations← Excel col 8 "[support: ...]" contents (new field)
  - metadata.audit_exempt     ← True iff col 19 contains B21 exempt marker
  - ground_truth.expected_response ← Excel col 11 (only if changed)

Non-destructive for other fields: key_facts, reasoning_chain, fail_conditions,
etc. are preserved as-is.

Design:
  - Loads Excel row map keyed by test_id.
  - Walks each JSONL file; for each line, if test_id is in the Excel map and
    any target field differs, rewrites the line.
  - --dry-run prints per-file change count without writing.
  - Creates .bak copies of JSONL files before overwriting (unless --no-backup).
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
EXCEL_PATH = (
    REPO_ROOT
    / "ground-truth"
    / "expert-validation"
    / "CCoP_V2_Test_Cases_Expert_Review.xlsx"
)
TEST_SUITE_DIR = REPO_ROOT / "ground-truth" / "test-suite"
SHEET_NAME = "Test Cases Review"

COL_TEST_ID = 1
COL_SECTION = 7
COL_CLAUSE_REFS = 8
COL_EXPECTED_RESP = 11
COL_REMARKS = 19

B21_EXEMPT_TOKEN = "[AUDIT_EXEMPT:"

_SUPPORT_RE = re.compile(r"\[support:\s*([^\]]+)\]", re.IGNORECASE)


def _parse_clause_refs(cell_value: str) -> tuple[list[str], list[str]]:
    """Parse col-8 text into (primary_clauses, support_citations).

    Input examples:
      "3.2.2(b), 3.2.2(c) [support: Risk Assessment Guide §4.2, §4.3]"
      "1.6.1, 1.6.2, 1.6.3 [support: Cybersecurity Act 2018 §11(7)]"
      "7.1.1(b), 7.1.1(g), 7.1.1(h) [support: 8.2.1]"
      "4.2.1"                         (legacy, unchanged rows)
      "8.2,8.4"                       (legacy B24, no space)
      ""                              (empty)
    """
    raw = (cell_value or "").strip()
    if not raw:
        return [], []
    # Extract support block first
    support: list[str] = []
    support_match = _SUPPORT_RE.search(raw)
    if support_match:
        support_text = support_match.group(1).strip()
        # Split on commas but don't split on commas inside parentheses
        # For simplicity we split on comma and strip — our patcher format
        # uses simple "A, B, C" within support brackets.
        support = [s.strip() for s in support_text.split(",") if s.strip()]
        raw = _SUPPORT_RE.sub("", raw).strip()
    # Now split remaining primary clauses
    primary_parts = [p.strip() for p in raw.split(",")]
    primary = [p for p in primary_parts if p]
    return primary, support


def _load_excel_map(
    excel_path: Path,
) -> dict[str, dict[str, Any]]:
    wb = load_workbook(excel_path, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{SHEET_NAME}' not found in {excel_path}. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]
    result: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        test_id = row[COL_TEST_ID - 1]
        if not test_id:
            continue
        section = row[COL_SECTION - 1]
        clause_cell = row[COL_CLAUSE_REFS - 1]
        expected_resp = row[COL_EXPECTED_RESP - 1]
        remarks = row[COL_REMARKS - 1]

        primary, support = _parse_clause_refs(
            clause_cell if isinstance(clause_cell, str) else str(clause_cell or "")
        )
        audit_exempt = B21_EXEMPT_TOKEN in str(remarks or "")
        result[str(test_id)] = {
            "section": str(section) if section is not None else "",
            "clause_reference": primary,
            "support_citations": support,
            "expected_response": str(expected_resp) if expected_resp is not None else "",
            "audit_exempt": audit_exempt,
        }
    return result


def _sync_line(
    line_obj: dict[str, Any],
    excel_entry: dict[str, Any],
) -> tuple[dict[str, Any], list[str]]:
    """Apply excel_entry to line_obj in-place-ish. Return (updated_obj, changes)."""
    changes: list[str] = []
    meta = line_obj.setdefault("metadata", {})

    new_section = excel_entry["section"]
    if new_section and str(meta.get("section", "")) != new_section:
        changes.append(f"section: '{meta.get('section')}' → '{new_section}'")
        meta["section"] = new_section

    new_clause = excel_entry["clause_reference"]
    # Compare list equality (order-insensitive is risky; keep Excel order)
    if new_clause and list(meta.get("clause_reference") or []) != new_clause:
        changes.append(
            f"clause_reference: {meta.get('clause_reference')} → {new_clause}"
        )
        meta["clause_reference"] = new_clause

    new_support = excel_entry["support_citations"]
    if new_support and list(meta.get("support_citations") or []) != new_support:
        changes.append(
            f"support_citations: {meta.get('support_citations')} → {new_support}"
        )
        meta["support_citations"] = new_support
    elif not new_support and "support_citations" in meta and meta["support_citations"]:
        # Only clear if Excel says no support AND there was previously one
        pass  # Keep existing; clearing is unsafe without explicit intent

    if excel_entry["audit_exempt"] and not meta.get("audit_exempt"):
        changes.append("audit_exempt: False → True")
        meta["audit_exempt"] = True

    new_resp = excel_entry["expected_response"]
    gt = line_obj.setdefault("ground_truth", {})
    if new_resp and gt.get("expected_response", "") != new_resp:
        changes.append("expected_response: updated")
        gt["expected_response"] = new_resp

    return line_obj, changes


def _backup(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    dst = path.with_suffix(path.suffix + f".{stamp}.bak")
    shutil.copy2(path, dst)
    return dst


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Preview only.")
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Skip creating .bak copies of JSONL files before overwrite.",
    )
    parser.add_argument(
        "--benchmark",
        action="append",
        help="Restrict to benchmark(s), e.g. B08, B22. Repeatable.",
    )
    parser.add_argument("--excel", type=Path, default=EXCEL_PATH)
    parser.add_argument("--test-suite", type=Path, default=TEST_SUITE_DIR)
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERROR: Excel not found at {args.excel}", file=sys.stderr)
        return 2
    if not args.test_suite.is_dir():
        print(f"ERROR: test-suite dir not found: {args.test_suite}", file=sys.stderr)
        return 3

    print(f"Loading Excel: {args.excel}")
    excel_map = _load_excel_map(args.excel)
    print(f"  {len(excel_map)} test cases loaded")

    jsonl_files = sorted(args.test_suite.glob("*.jsonl"))
    if args.benchmark:
        wanted = {b.lower() for b in args.benchmark}
        jsonl_files = [
            f for f in jsonl_files
            if any(f.stem.startswith(f"{w}_") for w in wanted)
        ]
    if not jsonl_files:
        print("No matching JSONL files.", file=sys.stderr)
        return 4

    grand_modified = 0
    grand_unchanged = 0
    grand_missing = 0
    for jpath in jsonl_files:
        lines_out: list[str] = []
        modified = 0
        unchanged = 0
        missing = 0
        per_line_changes: list[tuple[str, list[str]]] = []
        with jpath.open("r") as fh:
            for raw_line in fh:
                raw_line = raw_line.rstrip("\n")
                if not raw_line.strip():
                    lines_out.append(raw_line)
                    continue
                obj = json.loads(raw_line)
                tid = obj.get("test_id", "")
                entry = excel_map.get(tid)
                if entry is None:
                    missing += 1
                    lines_out.append(raw_line)
                    continue
                updated_obj, changes = _sync_line(obj, entry)
                if changes:
                    modified += 1
                    per_line_changes.append((tid, changes))
                    lines_out.append(json.dumps(updated_obj, ensure_ascii=False))
                else:
                    unchanged += 1
                    lines_out.append(raw_line)

        print(f"\n=== {jpath.name} ===")
        print(f"  modified: {modified}  unchanged: {unchanged}  missing-in-excel: {missing}")
        if per_line_changes:
            for tid, chs in per_line_changes[:3]:
                print(f"  e.g. {tid}:")
                for ch in chs:
                    print(f"    - {ch}")
            if len(per_line_changes) > 3:
                print(f"  ...and {len(per_line_changes) - 3} more")

        if modified > 0 and not args.dry_run:
            if not args.no_backup:
                bk = _backup(jpath)
                print(f"  backup: {bk.name}")
            with jpath.open("w") as fh:
                fh.write("\n".join(lines_out) + "\n")
            print(f"  wrote: {jpath}")

        grand_modified += modified
        grand_unchanged += unchanged
        grand_missing += missing

    print("\n" + "=" * 60)
    print(f"TOTAL modified: {grand_modified}")
    print(f"TOTAL unchanged: {grand_unchanged}")
    print(f"TOTAL missing-in-excel: {grand_missing}")
    print(f"DRY RUN: {args.dry_run}")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
