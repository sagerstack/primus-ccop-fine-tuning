#!/usr/bin/env python3
"""
Generate Excel spreadsheet for expert validation of v2 test cases.
Reads all v2 JSONL files from ground-truth/test-suite/ and creates
a comprehensive Excel workbook for domain expert review.
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def main():
    # Paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    test_suite_dir = os.path.join(script_dir, "../test-suite")
    output_file = os.path.join(script_dir, "CCoP_V2_Test_Cases_Expert_Review.xlsx")

    # Read all test cases
    all_cases = []
    for filename in sorted(os.listdir(test_suite_dir)):
        if filename.endswith('.jsonl'):
            benchmark = filename.replace('.jsonl', '').upper()
            filepath = os.path.join(test_suite_dir, filename)
            with open(filepath, 'r') as f:
                for line in f:
                    case = json.loads(line)
                    case['benchmark'] = benchmark
                    all_cases.append(case)

    print(f"Read {len(all_cases)} test cases from {len(set(c['benchmark'] for c in all_cases))} benchmarks")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases Review"

    # Define columns
    columns = [
        "Test ID", "Benchmark", "Sector", "Domain", "Difficulty", "Category",
        "CCoP Section", "Clause Refs", "Question", "Expected Label",
        "Expected Response", "Key Facts (Critical)", "Key Facts (Important/Supporting)",
        "Reasoning Chain", "Forbidden Claims", "Approved (Y/N)",
        "Accuracy (1-5)", "Completeness (1-5)", "Remarks"
    ]

    # Header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        cell.border = thin_border

    # Set column widths
    column_widths = {
        "Test ID": 12, "Benchmark": 10, "Sector": 12, "Domain": 10, "Difficulty": 10,
        "Category": 18, "CCoP Section": 12, "Clause Refs": 20, "Question": 40,
        "Expected Label": 15, "Expected Response": 50, "Key Facts (Critical)": 40,
        "Key Facts (Important/Supporting)": 40, "Reasoning Chain": 50, "Forbidden Claims": 40,
        "Approved (Y/N)": 12, "Accuracy (1-5)": 12, "Completeness (1-5)": 15, "Remarks": 30
    }

    for col_idx, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = column_widths.get(col_name, 15)

    # Data rows
    for row_idx, case in enumerate(all_cases, 2):
        test_id = case.get('test_id', '')
        benchmark = case.get('benchmark', '')

        # Extract sector
        input_data = case.get('input', {})
        metadata = case.get('metadata', {})

        # Sector detection
        sector = metadata.get('scenario_sector', 'General')
        if not sector or sector == 'Unknown':
            input_text = str(input_data).lower()
            sector_map = {
                'bank': 'Banking', 'financial': 'Banking',
                'health': 'Healthcare', 'hospital': 'Healthcare', 'medical': 'Healthcare',
                'energy': 'Energy', 'power': 'Energy',
                'transport': 'Transportation', 'aviation': 'Transportation', 'port': 'Transportation',
                'water': 'Water',
                'government': 'Government', 'govt': 'Government', 'agency': 'Government',
                'telecom': 'Telecommunications', 'telco': 'Telecommunications'
            }
            for keyword, sec in sector_map.items():
                if keyword in input_text:
                    sector = sec
                    break

        # Domain
        domain = metadata.get('domain', 'Unknown')

        # Difficulty
        difficulty = metadata.get('difficulty', 'medium')

        # Category
        category = metadata.get('scenario_type', metadata.get('test_category', ''))

        # Section and clause
        section = metadata.get('section', '')
        clause_raw = metadata.get('clause_reference', '')
        if isinstance(clause_raw, list):
            clause = ', '.join(str(c) for c in clause_raw)
        else:
            clause = str(clause_raw) if clause_raw else ''

        # Question
        question = input_data.get('question', '')

        # Ground truth
        ground_truth = case.get('ground_truth', {})
        expected_label = ground_truth.get('expected_label', '')
        expected_response = ground_truth.get('expected_response', '')

        # Key facts
        key_facts = ground_truth.get('key_facts', [])
        critical_facts = [f.get('fact', '') for f in key_facts if f.get('tier') == 'critical']
        other_facts = [f.get('fact', '') for f in key_facts if f.get('tier') != 'critical']

        critical_text = '; '.join(critical_facts) if critical_facts else ''
        other_text = '; '.join(other_facts) if other_facts else ''

        # Reasoning chain
        reasoning_raw = ground_truth.get('reasoning_chain', '')
        if isinstance(reasoning_raw, list):
            reasoning = '; '.join(str(r) for r in reasoning_raw)
        else:
            reasoning = str(reasoning_raw) if reasoning_raw else ''

        # Fail conditions
        fail_conditions = case.get('fail_conditions', {})
        forbidden = fail_conditions.get('incorrect_claims', []) + fail_conditions.get('hallucination_patterns', [])
        forbidden_text = '; '.join(forbidden) if forbidden else ''

        # Write row
        ws.cell(row=row_idx, column=1, value=test_id)
        ws.cell(row=row_idx, column=2, value=benchmark)
        ws.cell(row=row_idx, column=3, value=sector)
        ws.cell(row=row_idx, column=4, value=domain)
        ws.cell(row=row_idx, column=5, value=difficulty)
        ws.cell(row=row_idx, column=6, value=category)
        ws.cell(row=row_idx, column=7, value=section)
        ws.cell(row=row_idx, column=8, value=clause)
        ws.cell(row=row_idx, column=9, value=question)
        ws.cell(row=row_idx, column=10, value=expected_label)
        ws.cell(row=row_idx, column=11, value=expected_response)
        ws.cell(row=row_idx, column=12, value=critical_text)
        ws.cell(row=row_idx, column=13, value=other_text)
        ws.cell(row=row_idx, column=14, value=reasoning)
        ws.cell(row=row_idx, column=15, value=forbidden_text)

        # Review columns (empty for expert to fill)
        for col_idx in [16, 17, 18, 19]:
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save
    wb.save(output_file)
    print(f"Excel spreadsheet saved to {output_file}")
    print(f"Total rows: {len(all_cases) + 1}")

    # Print summary
    print("\nSummary by Benchmark:")
    benchmark_counts = {}
    for case in all_cases:
        bm = case['benchmark']
        benchmark_counts[bm] = benchmark_counts.get(bm, 0) + 1
    for bm in sorted(benchmark_counts.keys()):
        print(f"  {bm}: {benchmark_counts[bm]} cases")

if __name__ == "__main__":
    main()
