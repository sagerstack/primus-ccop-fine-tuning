#!/usr/bin/env python3
"""
Generate Excel spreadsheet for expert validation of test cases.
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Benchmark descriptions from benchmarks-updated.md
BENCHMARK_DESCRIPTIONS = {
    "B1": "Evaluates understanding of CII/CIIO scope, digital boundary, and applicability under the Cybersecurity Act - High Impact",
    "B2": "Learns audit-style compliance judgement once applicability is established - High Impact",
    "B3": "Evaluates nuanced conditional reasoning common in audits - High Impact",
    "B4": "Baseline knowledge check for CCoP structure and control coverage - High Impact",
    "B5": "Evaluates accurate paraphrasing and literal understanding of CCoP control requirements - Medium Impact",
    "B6": "Evaluates understanding beyond literal wording - High Impact",
    "B7": "Learns common compliance failure patterns - High Impact",
    "B8": "Encodes risk-based prioritisation logic - High Impact",
    "B9": "Improves recognition of compliance-specific risks - High Impact",
    "B10": "Structured risk explanation, scored via expert rubric - High Impact",
    "B11": "Learns proportional judgement of severity - High Impact",
    "B12": "Encodes CSA-style audit reasoning - High Impact",
    "B13": "Learns typical audit evidence expectations - High Impact",
    "B14": "Learns practical, proportionate remediation - High Impact",
    "B15": "Filters unrealistic advice - High Impact",
    "B16": "Evaluates post-control reasoning - High Impact",
    "B17": "Distinguishes documented policy from operational reality - Medium Impact",
    "B18": "Evaluates understanding of CIIO, CSA, Commissioner roles - Medium Impact",
    "B19": "Tests reasoning stability - Medium Impact",
    "B20": "Lightweight grounding sanity check - Low Impact",
    "B21": "Detects non-existent regulatory claims - Low Impact"
}

def format_list(items):
    """Format a list as bullet points."""
    if not items:
        return ""
    if isinstance(items, list):
        return "\n".join(f"• {item}" for item in items)
    return str(items)

def format_dict(d):
    """Format a dictionary as key: value pairs."""
    if not d:
        return ""
    if isinstance(d, dict):
        lines = []
        for k, v in d.items():
            if isinstance(v, (list, dict)):
                lines.append(f"{k}: {json.dumps(v, ensure_ascii=False)}")
            else:
                lines.append(f"{k}: {v}")
        return "\n".join(lines)
    return str(d)

def extract_benchmark_id(test_id):
    """Extract benchmark ID from test_id (e.g., B1-001 -> B1)."""
    return test_id.split('-')[0]

def load_test_cases(test_suite_dir):
    """Load all test cases from JSONL files."""
    test_cases = []

    for jsonl_file in sorted(test_suite_dir.glob("b*.jsonl")):
        with open(jsonl_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    test_case = json.loads(line)
                    test_cases.append(test_case)
                except json.JSONDecodeError as e:
                    print(f"Error parsing {jsonl_file.name}: {e}")

    return test_cases

def create_excel(test_cases, output_path):
    """Create Excel spreadsheet for expert review."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases - Expert Review"

    # Define headers
    headers = [
        "Test ID",
        "Benchmark",
        "Category",
        "Difficulty",
        "Domain",
        "What is Evaluated & Impact",
        "Section",
        "Clause Reference",
        "Related Clauses",
        "Question",
        "Expected Response",
        "Key Facts",
        "Expected Label",
        "Reasoning Dimensions",
        "Safety Checks",
        "Approved (Y/N)",
        "Remarks"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths
    column_widths = {
        1: 12,   # Test ID
        2: 35,   # Benchmark
        3: 15,   # Category
        4: 12,   # Difficulty
        5: 10,   # Domain
        6: 60,   # What is Evaluated & Impact
        7: 30,   # Section
        8: 18,   # Clause Reference
        9: 18,   # Related Clauses
        10: 60,  # Question
        11: 80,  # Expected Response
        12: 60,  # Key Facts
        13: 40,  # Expected Label
        14: 60,  # Reasoning Dimensions
        15: 40,  # Safety Checks
        16: 15,  # Approved (Y/N)
        17: 60   # Remarks
    }

    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Populate test cases
    for row_num, test_case in enumerate(test_cases, 2):
        benchmark_id = extract_benchmark_id(test_case.get('test_id', ''))

        # Extract related clauses from metadata
        metadata = test_case.get('metadata', {})
        related_sections = metadata.get('related_sections', [])
        related_clauses = ', '.join(related_sections) if related_sections else ''

        row_data = [
            test_case.get('test_id', ''),
            test_case.get('benchmark_type', ''),
            test_case.get('benchmark_category', ''),
            test_case.get('difficulty', ''),
            metadata.get('domain', ''),
            BENCHMARK_DESCRIPTIONS.get(benchmark_id, ''),
            test_case.get('section', ''),
            test_case.get('clause_reference', ''),
            related_clauses,
            test_case.get('question', ''),
            test_case.get('expected_response', ''),
            format_list(test_case.get('key_facts', [])),
            test_case.get('expected_label', ''),
            format_dict(test_case.get('reasoning_dimensions', {})),
            format_list(test_case.get('safety_checks', [])),
            '',  # Approved (Y/N) - empty for expert to fill
            ''   # Remarks - empty for expert to fill
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            # Convert any non-string values to string to avoid Excel errors
            if isinstance(value, (dict, list)):
                cell.value = json.dumps(value, ensure_ascii=False)
            else:
                cell.value = str(value) if value is not None else ''
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Set row height for readability
            ws.row_dimensions[row_num].height = 60

    # Add data validation for Approved column (column 16)
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv.error = 'Please select Y or N'
    dv.errorTitle = 'Invalid Entry'
    ws.add_data_validation(dv)
    dv.add(f'P2:P{len(test_cases) + 1}')  # Column P is the 16th column

    # Freeze header row and first 3 columns
    ws.freeze_panes = 'D2'

    # Save workbook
    wb.save(output_path)
    print(f"✅ Excel file created: {output_path}")
    print(f"   Total test cases: {len(test_cases)}")

def main():
    test_suite_dir = Path("/Users/sagarpratapsingh/dev/sagerstack/studio-ssdlc/ground-truth/phase-2/test-suite")
    output_path = Path("/Users/sagarpratapsingh/dev/sagerstack/studio-ssdlc/ground-truth/phase-2/expert-validation/CCoP_Test_Cases_Expert_Review.xlsx")

    print("Loading test cases...")
    test_cases = load_test_cases(test_suite_dir)

    print(f"Loaded {len(test_cases)} test cases")
    print("Creating Excel spreadsheet...")

    create_excel(test_cases, output_path)

    print("\n" + "="*60)
    print("EXPERT REVIEW EXCEL GENERATED")
    print("="*60)
    print(f"File: {output_path.name}")
    print(f"Location: {output_path.parent}")
    print(f"Test Cases: {len(test_cases)}")
    print("="*60)

if __name__ == "__main__":
    main()
